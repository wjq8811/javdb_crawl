#!/usr/bin/python3  
# -*- coding: utf-8 -*- 
import os
import json
import datetime
import openpyxl

def main(file_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "sheet1"
    sheet1 = wb[wb.sheetnames[0]]
    
    
    list1_1 = ['演员','番号','时间','标题','时间','评分','类型','海报','预告片','截图','链接信息','链接时间','中文字幕','链接地址']
    sheet1.append(list1_1)

    now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    # excel_name = file_path + '\\' + str(now) +'.xlsx'
    excel_name = r'C:\Users\w\Desktop\download_javdb\\' + str(now) +'.xlsx'
    wb.save(filename=excel_name)

    artor_list = []
    num = 0
    for artor in os.listdir(file_path):
        artor_path = os.path.join(file_path,artor)
        if os.path.isdir(artor_path):
            for fanhao in os.listdir(artor_path):
                fanhao_path = os.path.join(artor_path,fanhao)
                if os.path.isdir(fanhao_path):
                    json_path = os.path.join(fanhao_path,fanhao+'.json')
                    print(json_path)
                    try:
                       with open(json_path,'r', encoding='utf-8') as load_f:
                        load_dict = json.load(load_f)

                        title = ''.join(load_dict['title'])
                        fanhao = ''.join(load_dict['fanhao'])
                        time = ''.join(load_dict['time'])
                        scoring = ''.join(load_dict['scoring'])#评分 
                        type_ = '、'.join(load_dict['type_'])
                        performer = '、'.join(load_dict['performer'])#演员
                        poster = ''.join(load_dict['poster'])#海报
                        trailer = ''.join(load_dict['trailer'])#预告片
                        runtimeg = ''.join(load_dict['runtimeg'])#片长
                        #截图列表'images_list':images_list
                        images = '|'.join(load_dict['images_list'])

                        for magnets in load_dict['magnet_list']:
                            magnet_link,magnet_time,magnet_info = magnets

                            #是否中文
                            is_zh = '否'
                            if ('字幕' in magnet_info) or ('中文' in magnet_info):
                                is_zh = '是'
                                break
                        if is_zh == '否':
                            magnet_link,magnet_time,magnet_info = load_dict['magnet_list'][0]


                        fanhao_list = [performer,fanhao,time,title,runtimeg,scoring,type_,poster,trailer,images,magnet_info,magnet_time,is_zh,magnet_link]
                        artor_list.append(fanhao_list)
                        print(performer,fanhao)
                        print('_'*20)
                    except Exception as e:
                        continue


                    # for x in tmp_list:
                    #     print(x.encode('gbk', 'ignore').decode('gbk'))
            num +=1
            if num == 50:
                print('num:',num)
                wb = openpyxl.load_workbook(excel_name)
                sheet1 = wb[wb.sheetnames[0]]
                for fanhao_list_ in artor_list:
                    sheet1.append(fanhao_list_)
                wb.save(filename=excel_name)
                num = 0 
                artor_list = []

    wb = openpyxl.load_workbook(excel_name)
    sheet1 = wb[wb.sheetnames[0]]
    for fanhao_list_ in artor_list:
        sheet1.append(fanhao_list_)
    wb.save(filename=excel_name)
    
    print('文件保存在'+ excel_name)



if __name__ == '__main__':
    path = r'Z:\censored'
    main(path)